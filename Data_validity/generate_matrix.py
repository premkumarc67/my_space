"""
Four-Matrix Column Audit Generator
-------------------------------------
Produces a single Excel workbook with four sheets:

  Sheet 1 – Column Presence  : Is the column present in the file?          YES / NO
  Sheet 2 – Is Blank         : Is the column entirely null (no values)?     YES / NO / N/A
  Sheet 3 – Zero Count       : How many cells in the column equal exactly 0? integer / N/A
  Sheet 4 – Full Coverage    : % of rows that are non-null (zeros count as filled) float% / N/A

Rows    → expected column names (headers of col_names.xlsx)
Columns → one per data file

Color logic:
  Sheet 1 – Presence  : YES=green,      NO=red,                N/A=grey
  Sheet 2 – Is Blank  : YES(blank)=red, NO(has values)=green,  N/A=grey
  Sheet 3 – Zero Count: 0=green,  1-10=yellow,  11-50=orange,  >50=red,  N/A=grey
  Sheet 4 – Coverage  : 100%=green,  95-99%=light-green,  80-94%=yellow,
                        50-79%=orange,  <50%=red,  N/A=grey
  Header row          : dark navy, white bold text

Usage:
    python generate_matrix.py

Requirements:
    pip install pandas openpyxl
"""

import os
import glob
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# CONFIG
COL_NAMES_FILE  = "col_names.xlsx"
DATA_FILES_GLOB = "*.xlsx"
EXCLUDE_FILES   = {"col_names.xlsx", "example.xlsx", "column_matrix.xlsx"}
OUTPUT_FILE     = "column_matrix.xlsx"

# COLOUR PALETTE
HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"

# Semantic colours
GREEN_S    = ("C6EFCE", "1E6B2E")   # (bg, fg)  — good
LGREEN_S   = ("E2EFDA", "375623")   # light green — near good
YELLOW_S   = ("FFEB9C", "7B5B00")   # caution
ORANGE_S   = ("FFD2A0", "7B3F00")   # warning
RED_S      = ("FFCCCC", "9C0006")   # bad
GREY_S     = ("F2F2F2", "808080")   # N/A / absent


ROW_ALT    = "FAFAFA"
ROW_NORM   = "FFFFFF"



def _fill(bg):
    return PatternFill("solid", fgColor=bg)

def _font(fg, bold=False):
    return Font(name="Arial", size=10, color=fg, bold=bold)


# ── Sheet 1: Presence colour (YES/NO) ──────────────────────────────
def presence_colour(value, is_summary=False):
    bg, fg = {
        "YES": GREEN_S,
        "NO":  RED_S,
    }.get(value, GREY_S)
    return bg, fg, False


# ── Sheet 2: Is Blank colour (YES/NO) ──────────────────────────────
def blank_colour(value, is_summary=False):
    # YES means blank = bad (red); NO means has data = good (green)
    bg, fg = {
        "YES": RED_S,
        "NO":  GREEN_S,
    }.get(value, GREY_S)
    return bg, fg, False


# ── Sheet 3: Zero count colour ─────────────────────────────────────
def zero_colour(value, is_summary=False):
    """value is an int (zero count) or 'N/A'."""
    if value == "N/A":
        bg, fg = GREY_S
    else:
        try:
            n = float(value)
        except (ValueError, TypeError):
            bg, fg = GREY_S
        else:
            if n == 0:
                bg, fg = GREEN_S
            elif n <= 10:
                bg, fg = YELLOW_S
            elif n <= 50:
                bg, fg = ORANGE_S
            else:
                bg, fg = RED_S
    return bg, fg, False


# ── Sheet 4: Coverage colour ────────────────────────────────────────
def coverage_colour(value, is_summary=False):
    """value is a float 0–100 or 'N/A'."""
    if value == "N/A":
        bg, fg = GREY_S
    else:
        try:
            pct = float(str(value).replace("%", ""))
        except (ValueError, TypeError):
            bg, fg = GREY_S
        else:
            if pct >= 100:
                bg, fg = GREEN_S
            elif pct >= 95:
                bg, fg = LGREEN_S
            elif pct >= 80:
                bg, fg = YELLOW_S
            elif pct >= 50:
                bg, fg = ORANGE_S
            else:
                bg, fg = RED_S
    return bg, fg, False


COLOUR_FN = {
    "1. Column Presence": presence_colour,
    "2. Is Blank":        blank_colour,
    "3. Zero Count":      zero_colour,
    "4. Full Coverage":   coverage_colour,
}



# DATA ANALYSIS

def get_row_names(col_names_file):
    return list(pd.read_excel(col_names_file, nrows=0).columns)


def analyse_file(filepath, row_names):
    df        = pd.read_excel(filepath)
    file_cols = set(df.columns)
    total     = len(df)

    presence, is_blank, zero_count, coverage = {}, {}, {}, {}

    for col in row_names:
        if col not in file_cols:
            presence[col]   = "NO"
            is_blank[col]   = "N/A"
            zero_count[col] = "N/A"
            coverage[col]   = "N/A"
        else:
            series = df[col]
            presence[col]   = "YES"
            is_blank[col]   = "YES" if series.isna().all() else "NO"

            try:
                zeros = int((series == 0).sum())
            except TypeError:
                zeros = 0
            zero_count[col] = zeros

            filled = int(series.notna().sum())
            coverage[col]   = round(filled / total * 100, 1) if total > 0 else 0.0

    return {
        "presence":   presence,
        "is_blank":   is_blank,
        "zero_count": zero_count,
        "coverage":   coverage,
    }


def build_matrices(row_names, data_files):
    recs = {"presence": {}, "is_blank": {}, "zero_count": {}, "coverage": {}}

    for fpath in data_files:
        fname  = os.path.basename(fpath)
        result = analyse_file(fpath, row_names)
        for key in recs:
            recs[key][fname] = result[key]

    def make_df(key):
        df = pd.DataFrame(recs[key], index=row_names)
        df.index.name = "Column Name"
        return df.reset_index()

    df_pres  = make_df("presence")
    df_blank = make_df("is_blank")
    df_zero  = make_df("zero_count")
    df_cov   = make_df("coverage")

    return df_pres, df_blank, df_zero, df_cov



# STYLING
def style_sheet(ws, sheet_name):
    colour_fn = COLOUR_FN[sheet_name]
    thin      = Side(border_style="thin", color="D0D0D0")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    max_col = ws.max_column
    max_row = ws.max_row

    # Header
    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = _fill(HEADER_BG), _font(HEADER_FG, bold=True)
        cell.border    = border
        cell.alignment = left_wrap if c == 1 else center

    # Body
    for r in range(2, max_row + 1):
        alt_bg = ROW_ALT if r % 2 == 0 else ROW_NORM

        for c in range(1, max_col + 1):
            cell  = ws.cell(r, c)
            value = cell.value
            cell.border = border

            if c == 1:                        # Column Name
                cell.fill      = _fill(alt_bg)
                cell.font      = _font("000000")
                cell.alignment = left_wrap

            else:                             # File columns
                bg, fg, bold = colour_fn(value, is_summary=False)
                cell.fill      = _fill(bg)
                cell.font      = _font(fg, bold=bold)
                cell.alignment = center
                if sheet_name == "4. Full Coverage" and value != "N/A":
                    try:
                        cell.number_format = '0.0"%"'
                    except Exception:
                        pass

    # Widths
    ws.column_dimensions["A"].width = 70
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # Heights
    ws.row_dimensions[1].height = 22
    for r in range(2, max_row + 1):
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "B2"
    ws.title = sheet_name


def write_workbook(df_pres, df_blank, df_zero, df_cov, output_file):
    sheets = [
        ("1. Column Presence", df_pres),
        ("2. Is Blank",        df_blank),
        ("3. Zero Count",      df_zero),
        ("4. Full Coverage",   df_cov),
    ]

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for name, df in sheets:
            df.to_excel(writer, sheet_name=name, index=False)

    wb = load_workbook(output_file)
    for name, _ in sheets:
        style_sheet(wb[name], name)

    wb.save(output_file)
    print(f"✅  Workbook saved → {output_file}")



# MAIN

def main():
    row_names = get_row_names(COL_NAMES_FILE)
    print(f"Loaded {len(row_names)} expected columns from '{COL_NAMES_FILE}'")

    data_files = sorted(
        f for f in glob.glob(DATA_FILES_GLOB)
        if os.path.basename(f) not in EXCLUDE_FILES
    )
    if not data_files:
        raise FileNotFoundError("No data files found. Check DATA_FILES_GLOB / EXCLUDE_FILES.")
    print(f"Found {len(data_files)} data file(s): {[os.path.basename(f) for f in data_files]}")

    df_pres, df_blank, df_zero, df_cov = build_matrices(row_names, data_files)
    write_workbook(df_pres, df_blank, df_zero, df_cov, OUTPUT_FILE)


if __name__ == "__main__":
    main()